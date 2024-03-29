# coding=utf-8
from pykeyboard import PyKeyboard
from pymouse import PyMouse
import time
import pyHook
import pythoncom
import xlrd
import xlwt
import pyperclip
from pynput import mouse, keyboard
import threading
import sys
import re
from openpyxl import Workbook, load_workbook


#-------------------

import requests
import urllib
import random

from urllib import request
import os
import random
import requests
#导入需要用到的模块
import requests

import os
from urllib.request import urlretrieve

import requests
import hashlib





def saveToExcel(name,link,isDownloaded=''):
    excel = xlrd.open_workbook(excelUrl)
    table = excel.sheets()[0]
    rowCount = table.nrows
    colCount = table.ncols


    wb= load_workbook(filename=excelUrl)
    worksheet=wb.active
    worksheet=wb['Sheet1']

    hash_str=Hash(name)

    worksheet.cell(row=rowCount+1,column=1,value=name)
    worksheet.cell(row=rowCount+1,column=2,value=link)
    worksheet.cell(row=rowCount+1,column=3,value=isDownloaded)
    worksheet.cell(row=rowCount+1,column=4,value=hash_str)

    wb.save(excelUrl)

    return hash_str




def CheckIfExist(check_name):
    excel = xlrd.open_workbook(excelUrl)
    table = excel.sheets()[0]
    rowCount = table.nrows
    colCount = table.ncols

    for i in range(rowCount):
        name = str(table.cell_value(i, 0))
        link = str(table.cell_value(i, 1))
        isDownloaded = str(table.cell_value(i, 2))
        if(name=='name'):
            continue
        if(name==check_name):
            return True


    return False
def CheckIfDownload(check_name):
    excel = xlrd.open_workbook(excelUrl)
    table = excel.sheets()[0]
    rowCount = table.nrows
    colCount = table.ncols


    for i in range(rowCount):
        name = str(table.cell_value(i, 0))
        link = str(table.cell_value(i, 1))
        isDownloaded = str(table.cell_value(i, 2))
        if(isDownloaded=='isDownloaded'):
            continue
        if(isDownloaded=='download'):
            return True


    return False




# saveToExcel('2','xxx.com','choose','')
# saveToExcel('3','xxx.com','choose','')
# saveToExcel('未完成','xxx.com','choose','')
# saveToExcel('test','xxx.com','choose','')

def Update_Img():
    for f in os.listdir(r'C:\Users\Administrator\Desktop\t'):#通过图片更新excel
        print(f)
        name=str(f).replace('.mp4','')
        video_link=''

        exist=CheckIfExist(name)
        if(not exist):
            saveToExcel(name,video_link)
        else:
            print('重复：'+name)

def Update_Video():
    for f in os.listdir(r'C:\Users\Administrator\Desktop\t'):#通过视频更新excel
        print(f)
        file_name=str(f).replace('.mp4','')

        isdownload=CheckIfDownload(file_name)
        if('未完成' in isdownload):
            isdownload=False

            #os.remove('')#下一半的，删除了

        if(isdownload==False):
            开始下载


def Update_Video(name):#下载好之后设置isDownload的值
    excel = xlrd.open_workbook(excelUrl)
    table = excel.sheets()[0]
    colCount = table.ncols

    rowCount = table.nrows
    cur_rowCount=0
    wb = load_workbook(filename=excelUrl)
    worksheet = wb.active
    worksheet = wb['Sheet1']

    for i in range(rowCount):
        excel_name = str(table.cell_value(i, 0))
        cur_rowCount+=1
        if (name == excel_name):
            worksheet.cell(row=cur_rowCount, column=3, value='√')

    wb.save(excelUrl)

    print('表格更新完成')


def SaveLastName(name):
    excel = xlrd.open_workbook(excelUrl)
    table2 = excel.sheets()[1]
    rowCount = table2.nrows
    colCount = table2.ncols

    wb = load_workbook(filename=excelUrl)
    worksheet = wb.active
    worksheet = wb['Sheet2']

    worksheet.cell(row=rowCount + 1, column=1, value=name)

    wb.save(excelUrl)

def GetLastName():
    excel = xlrd.open_workbook(excelUrl)
    table2 = excel.sheets()[1]
    rowCount = table2.nrows
    colCount = table2.ncols
    last_name = str(table2.cell_value(rowCount-1, 0))
    return last_name

def Hash(str):
    signaturemd5 = hashlib.md5()
    signaturemd5.update(str.encode('utf8'))
    hash_str = signaturemd5.hexdigest()
    return hash_str

def GetName_ByHash(hash_str):
    exce = xlrd.open_workbook(excelUrl)
    table = excel.sheets()[0]
    rowCount = table.nrows
    colCount = table.ncols
    for i in range(rowCount):
        excel_hash = str(table.cell_value(i, 3))
        name=str(table.cell_value(i, 1))
        if(hash_str==excel_hash):
            return name

    return ''
def GetColValues(excel_url,col,sheet=0):
    excel = xlrd.open_workbook(excel_url)
    table = excel.sheets()[sheet]
    rowCount = table.nrows
    colCount = table.ncols
    values=[]
    for r in range(rowCount):
        values.append(table.cell_value(r,col))
    return values

def GetRowValues(excel_url,row,sheet=0):
    excel = xlrd.open_workbook(excel_url)
    table = excel.sheets()[sheet]
    rowCount = table.nrows
    colCount = table.ncols
    values=[]
    for col in range(colCount):
        values.append(table.cell_value(row,col))
    return values

def WriteExcel(row,col,value,wb,url):



    worksheet = wb.active
    worksheet=wb._sheets[0]
    worksheet.cell(row=row, column=col, value=value)


def Do(target1, target2,excelUrl):


    target_count = 0
    target_money = 0

    features = GetColValues(excelUrl, 3)
    counts = GetColValues(excelUrl, 5)
    moneys = GetColValues(excelUrl, 7)
    names = GetColValues(excelUrl, 2)

    for index in range(len(counts)):
        feature = features[index]
        name = names[index]
        count = (counts[index])
        money = (moneys[index])

        isNumber = str(count).replace('.', '').isdigit()
        if (count == ''):
            continue
        if (money == ''):
            continue
        if('电力电缆头'==name):
            continue

        if (isNumber):
            if ('甲供' in name):
                continue
            if('甲供' in feature):
                continue
            count = float(count)

            if (name in target1):  # todo----------------

                if (target2 == '' or (target2 in feature)):
                    target_count += count
                    target_money += money
                    WriteExcel(index + 1, 15, target1 + '  /  ' + target2, wb, excelUrl)

        index = index + 1


    wb.save(excelUrl)

    target_list=[]
    # print('---------')
    # print(target1)
    # print(target2)
    # print(target_money)
    # print(target_count)

    target_list.append(target_count)
    target_list.append(target_money)
    return target_list

dir=r"E:\我\备份\安装\TODODODODDODO\4-29\水电"
dir=r"E:\我\备份\安装\TODODODODDODO\4-29\消防"
dir=r"E:\我\备份\安装\TODODODODDODO\4-29\弱电"

target_excelUrl = r"E:\我\备份\安装\TODODODODDODO\4-29\水电项目管理.xlsx"
target_excelUrl = r"E:\我\备份\安装\TODODODODDODO\4-29\消防项目管理.xlsx"
target_excelUrl = r"E:\我\备份\安装\TODODODODDODO\4-29\弱电项目管理.xlsx"
files=os.listdir(dir)
targets=[]
wb2=load_workbook(filename=target_excelUrl)
worksheet2=wb2.active
worksheet2 = wb2._sheets[0]

rows=GetRowValues(target_excelUrl,1)
for r in rows:
    targets.append(r)

print(targets)
target_row=4
for file in files:
    excelUrl = dir + '\\' + file
    print('加载表格！')
    wb = load_workbook(filename=excelUrl)
    worksheet2.cell(row=target_row, column=1, value=file)

    for i in range(len(targets)):
        if (i % 2 == 0):
            continue


        target1 = targets[i]
        target2 = targets[i + 1]

        if ('$' in excelUrl):
            continue
        targetList=Do(target1, target2, excelUrl)
        print(targetList)
        if(targetList==[0,0]):
            continue
        worksheet2.cell(row=target_row, column=i+1, value=targetList[0])
        worksheet2.cell(row=target_row, column=i+2, value=targetList[1])
    target_row+=1
    wb2.save(target_excelUrl)
